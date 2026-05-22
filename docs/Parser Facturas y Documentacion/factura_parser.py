#!/usr/bin/env python3
"""
FACTURA PARSER - DataFlow
=========================
Extrae datos de facturas (imágenes o PDF multipágina) usando Claude Vision
y genera un Excel estructurado listo para importar en DataFlow.

PDF: procesa **cada página** como un comprobante. El estado PAGADA / POR_PAGAR se infiere
del nombre del archivo (prefijo PAGADAS / POR PAGAR).

USO:
  python factura_parser.py --input ./facturas/ --output facturas_extraidas.xlsx
  python factura_parser.py --input ./facturas/ --output result.xlsx --delay 1.5
  python factura_parser.py --input ./facturas/ --resume-from 120

FORMATOS: JPG, JPEG, PNG, WEBP, PDF (todas las páginas)
"""

from __future__ import annotations

import argparse
import base64
import glob
import json
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import anthropic
except ImportError:
    print("ERROR: Instalar anthropic:  pip install anthropic")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Instalar openpyxl:  pip install openpyxl")
    sys.exit(1)

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False


PDF_RENDER_ZOOM = 1.35
# Anthropic rechaza payloads enormes; raster grande → 400 invalid_request
MAX_IMAGE_DIMENSION = 1800
CLAUDE_MAX_TOKENS = 8192
CLAUDE_MODEL = __import__("os").environ.get("ANTHROPIC_INVOICE_MODEL", "claude-opus-4-5")


EXTRACTION_PROMPT = """Analizá esta imagen de factura argentina y extraé los datos en formato JSON.

Respondé ÚNICAMENTE con el JSON, sin texto adicional, sin markdown, sin backticks.

Estructura exacta requerida:
{
  "cabecera": {
    "punto_venta": "string (ej: 0001)",
    "numero_comprobante": "string (ej: 00000123)",
    "fecha_emision": "string (ej: 15/03/2025)",
    "tipo_comprobante": "string (ej: FACTURA A, FACTURA B, REMITO, NOTA DE CREDITO A)",
    "proveedor_nombre_comercial": "string",
    "razon_social": "string",
    "cuit": "string (ej: 20-12345678-9)",
    "condicion_iva_proveedor": "string (ej: Responsable Inscripto, Monotributista, Exento)",
    "condicion_iva_receptor": "string si figura",
    "receptor_razon_social": "string si figura",
    "receptor_cuit": "string si figura"
  },
  "items": [
    {
      "descripcion": "string",
      "cantidad": "string (número con decimales si aplica)",
      "unidad_medida": "string si figura (ej: kg, lt, un)",
      "precio_unitario": "string (número)",
      "subtotal": "string (número)"
    }
  ],
  "impuestos": [
    {
      "nombre": "string (ej: IVA 21%, IVA 10.5%, Percepcion IIBB)",
      "base_imponible": "string (número o vacío si no figura)",
      "porcentaje": "string (ej: 21, 10.5 o vacío)",
      "importe": "string (número)"
    }
  ],
  "totales": {
    "subtotal_neto": "string (número, sin impuestos)",
    "descuento_importe": "string (número o 0)",
    "total_impuestos": "string (número)",
    "total_factura": "string (número total final)"
  },
  "notas": "string con cualquier observación relevante o campos que no pudiste leer con claridad"
}

Reglas importantes:
- Si un campo no está visible o no existe, usá "" (string vacío)
- Para números, usá punto decimal (ej: 1250.50), no comas
- precio_unitario y subtotal en cada item son IMPORTES NETOS SIN IVA (gravados en Factura A/B)
- El punto_venta suele estar antes del guión en el número de comprobante (ej: 0001-00000123)
- Extraé TODOS los items de la factura, no omitas ninguno
"""


def estado_pago_desde_nombre_archivo(stem: str) -> str:
    u = stem.upper().strip()
    if u.startswith("PAGADAS"):
        return "PAGADA"
    if u.startswith("POR PAGAR") or u.startswith("POR_PAGAR"):
        return "POR_PAGAR"
    return "DESCONOCIDO"


def strip_accents_lower(s: str) -> str:
    t = unicodedata.normalize("NFD", s.lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def normalizar_condicion_iva_dataflow(texto: str) -> tuple[str, str]:
    """Devuelve (texto OCR original limpio, código DataFlow)."""
    raw_original = (texto or "").strip()
    t = strip_accents_lower(raw_original)
    if not t:
        return raw_original, "responsable_inscripto"
    if "monotributo" in t:
        return raw_original, "monotributista"
    if "exento" in t:
        return raw_original, "exento"
    if "consumidor final" in t:
        return raw_original, "consumidor_final"
    if (
        "responsable inscripto" in t
        or "resp inscripto" in t
        or "resp.inscripto" in t
        or "responsable inscripto" in t.replace(".", "")
    ):
        return raw_original, "responsable_inscripto"
    return raw_original, "responsable_inscripto"


def enrich_invoice_record(data: dict) -> dict:
    cab = data.setdefault("cabecera", {})
    raw = cab.get("condicion_iva_proveedor", "") or ""
    disp, code = normalizar_condicion_iva_dataflow(raw)
    cab["condicion_iva_proveedor_display"] = disp
    cab["condicion_iva_proveedor_dataflow"] = code
    return data


def encode_image_file(image_path: str) -> tuple[str, str]:
    ext = Path(image_path).suffix.lower()
    media_type_map = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }
    media_type = media_type_map.get(ext, "image/jpeg")
    with open(image_path, "rb") as f:
        data = base64.standard_b64encode(f.read()).decode("utf-8")
    return data, media_type


def encode_pdf_page_as_png(pdf_path: str, page_index: int) -> tuple[str, str]:
    if not HAS_PYMUPDF:
        raise ImportError("Para PDF necesitás: pip install pymupdf")
    doc = fitz.open(pdf_path)
    try:
        page = doc[page_index]
        zoom = float(PDF_RENDER_ZOOM)
        pix = None
        for _ in range(14):
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            if pix.width <= MAX_IMAGE_DIMENSION and pix.height <= MAX_IMAGE_DIMENSION:
                break
            zoom *= 0.72
        assert pix is not None
        jpg = pix.tobytes("jpeg", jpg_quality=88)
        png = pix.tobytes("png")
        if len(jpg) <= len(png) * 1.05:
            return base64.standard_b64encode(jpg).decode("utf-8"), "image/jpeg"
        return base64.standard_b64encode(png).decode("utf-8"), "image/png"
    finally:
        doc.close()


def extract_invoice_data(
    client: anthropic.Anthropic,
    *,
    image_b64: str,
    media_type: str,
    meta: dict,
) -> dict:
    message = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=CLAUDE_MAX_TOKENS,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_b64,
                        },
                    },
                    {"type": "text", "text": EXTRACTION_PROMPT},
                ],
            }
        ],
    )

    response_text = message.content[0].text.strip()
    if response_text.startswith("```"):
        lines = response_text.split("\n")
        response_text = "\n".join(lines[1:-1])

    data = json.loads(response_text)
    enrich_invoice_record(data)
    data["_job_key"] = meta["job_key"]
    data["_archivo_pdf"] = meta["pdf_name"]
    data["_pagina_pdf"] = meta["page_display"]
    data["_estado_pago"] = meta["estado_pago"]
    data["_archivo"] = meta["vista_archivo"]
    data["_procesado"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    data["_error"] = ""
    return data


def _unique_paths(paths: list[str]) -> list[str]:
    """En Windows, *.pdf y *.PDF encuentran los mismos archivos y duplicaban el trabajo."""
    seen: set[str] = set()
    out: list[str] = []
    for p in paths:
        key = str(Path(p).resolve())
        if key not in seen:
            seen.add(key)
            out.append(p)
    return out


def get_input_files(input_path: str) -> list[str]:
    input_path = Path(input_path)
    if input_path.is_file():
        return [str(input_path)]
    if input_path.is_dir():
        extensions = [
            "*.jpg",
            "*.jpeg",
            "*.png",
            "*.webp",
            "*.pdf",
            "*.JPG",
            "*.JPEG",
            "*.PNG",
            "*.PDF",
        ]
        files: list[str] = []
        for ext in extensions:
            files.extend(glob.glob(str(input_path / ext)))
        files = _unique_paths(files)
        files.sort(key=lambda p: p.lower())
        return files
    files = glob.glob(input_path)
    files.sort()
    return files


def flatten_jobs(file_paths: list[str]) -> list[dict]:
    jobs: list[dict] = []
    for fp in file_paths:
        path = Path(fp)
        estado = estado_pago_desde_nombre_archivo(path.stem)
        if path.suffix.lower() == ".pdf":
            if not HAS_PYMUPDF:
                raise ImportError("Hay PDF pero falta pymupdf")
            doc = fitz.open(fp)
            try:
                n = len(doc)
                for i in range(n):
                    page_no = i + 1
                    job_key = f"{path.name}|{page_no}"
                    jobs.append(
                        {
                            "kind": "pdf",
                            "path": fp,
                            "page_index": i,
                            "job_key": job_key,
                            "estado_pago": estado,
                            "pdf_name": path.name,
                            "page_display": str(page_no),
                            "vista_archivo": f"{path.name} (pág. {page_no})",
                        }
                    )
            finally:
                doc.close()
        else:
            job_key = f"{path.name}|1"
            jobs.append(
                {
                    "kind": "img",
                    "path": fp,
                    "page_index": 0,
                    "job_key": job_key,
                    "estado_pago": estado,
                    "pdf_name": path.name,
                    "page_display": "1",
                    "vista_archivo": path.name,
                }
            )
    return jobs


HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ACCENT_FILL = PatternFill("solid", start_color="BDD7EE")
ERROR_FILL = PatternFill("solid", start_color="FFE0E0")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
ALT_FILL = PatternFill("solid", start_color="F2F7FC")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0C4DE"),
    right=Side(style="thin", color="B0C4DE"),
    top=Side(style="thin", color="B0C4DE"),
    bottom=Side(style="thin", color="B0C4DE"),
)


def style_header_cell(cell, text, size=11, bold=True, white=True):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, size=size, color="FFFFFF" if white else "1F4E79")
    cell.fill = HEADER_FILL if white else ACCENT_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def create_excel(all_data: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    ws_facturas = wb.active
    ws_facturas.title = "Facturas"
    ws_facturas.sheet_view.showGridLines = False
    ws_facturas.freeze_panes = "A3"

    last_col = "U"
    ws_facturas.merge_cells(f"A1:{last_col}1")
    title_cell = ws_facturas["A1"]
    title_cell.value = "DATAFLOW — IMPORTACIÓN DE FACTURAS (OCR)"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_facturas.row_dimensions[1].height = 30

    headers = [
        "Clave trabajo",
        "Estado Pago",
        "Archivo PDF",
        "Pagina",
        "Archivo vista",
        "Punto Venta",
        "Nro Comprobante",
        "Fecha Emisión",
        "Tipo Comprobante",
        "Nombre Comercial",
        "Razón Social",
        "CUIT",
        "Cond. IVA proveedor (Dataflow)",
        "Cond. IVA (texto OCR)",
        "Subtotal Neto",
        "Descuento",
        "Total Impuestos",
        "Total Factura",
        "Procesado",
        "Error",
        "Notas",
    ]
    col_widths = [28, 14, 28, 8, 36, 12, 14, 14, 18, 28, 36, 18, 28, 28, 14, 12, 14, 14, 18, 36, 36]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_facturas.cell(row=2, column=col_idx)
        style_header_cell(cell, header, size=10)
        ws_facturas.column_dimensions[get_column_letter(col_idx)].width = width
    ws_facturas.row_dimensions[2].height = 35

    for row_idx, data in enumerate(all_data, 3):
        alt = row_idx % 2 == 0
        cab = data.get("cabecera", {})
        tot = data.get("totales", {})
        error = data.get("_error", "")
        fill = ERROR_FILL if error else (ALT_FILL if alt else WHITE_FILL)

        df_iva = cab.get("condicion_iva_proveedor_dataflow", "") or cab.get(
            "condicion_iva_proveedor_normalizada", ""
        )
        row_data = [
            data.get("_job_key", ""),
            data.get("_estado_pago", ""),
            data.get("_archivo_pdf", ""),
            data.get("_pagina_pdf", ""),
            data.get("_archivo", ""),
            cab.get("punto_venta", ""),
            cab.get("numero_comprobante", ""),
            cab.get("fecha_emision", ""),
            cab.get("tipo_comprobante", ""),
            cab.get("proveedor_nombre_comercial", ""),
            cab.get("razon_social", ""),
            cab.get("cuit", ""),
            df_iva,
            cab.get("condicion_iva_proveedor_display", "") or cab.get("condicion_iva_proveedor", ""),
            tot.get("subtotal_neto", ""),
            tot.get("descuento_importe", ""),
            tot.get("total_impuestos", ""),
            tot.get("total_factura", ""),
            data.get("_procesado", ""),
            error,
            data.get("notas", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws_facturas.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws_facturas.row_dimensions[row_idx].height = 20

    ws_items = wb.create_sheet("Items de Facturas")
    ws_items.sheet_view.showGridLines = False
    ws_items.freeze_panes = "A3"
    ws_items.merge_cells("A1:L1")
    t2 = ws_items["A1"]
    t2.value = "DATAFLOW — DETALLE DE ÍTEMS POR FACTURA"
    t2.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t2.fill = HEADER_FILL
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_items.row_dimensions[1].height = 30

    item_headers = [
        "Clave trabajo",
        "Archivo vista",
        "Punto Venta",
        "Nro Comprobante",
        "Fecha",
        "Descripción Ítem",
        "Cantidad",
        "Unidad Medida",
        "Precio Unitario",
        "Subtotal Ítem",
        "Estado Pago",
    ]
    item_widths = [28, 36, 12, 14, 14, 45, 12, 14, 14, 14, 12]

    for col_idx, (header, width) in enumerate(zip(item_headers, item_widths), 1):
        cell = ws_items.cell(row=2, column=col_idx)
        style_header_cell(cell, header, size=10)
        ws_items.column_dimensions[get_column_letter(col_idx)].width = width
    ws_items.row_dimensions[2].height = 35

    item_row = 3
    for data in all_data:
        if data.get("_error"):
            continue
        cab = data.get("cabecera", {})
        items = data.get("items", [])
        estado = data.get("_estado_pago", "")
        job_key = data.get("_job_key", "")
        vista = data.get("_archivo", "")
        for i, item in enumerate(items):
            alt = (item_row + i) % 2 == 0
            row_data = [
                job_key,
                vista,
                cab.get("punto_venta", ""),
                cab.get("numero_comprobante", ""),
                cab.get("fecha_emision", ""),
                item.get("descripcion", ""),
                item.get("cantidad", ""),
                item.get("unidad_medida", ""),
                item.get("precio_unitario", ""),
                item.get("subtotal", ""),
                estado,
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_items.cell(row=item_row + i, column=col_idx)
                cell.value = value
                cell.font = Font(name="Arial", size=10)
                cell.fill = ALT_FILL if alt else WHITE_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = THIN_BORDER
            ws_items.row_dimensions[item_row + i].height = 18
        item_row += max(len(items), 1)

    ws_imp = wb.create_sheet("Impuestos")
    ws_imp.sheet_view.showGridLines = False
    ws_imp.freeze_panes = "A3"
    ws_imp.merge_cells("A1:I1")
    t3 = ws_imp["A1"]
    t3.value = "DATAFLOW — IMPUESTOS POR FACTURA"
    t3.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t3.fill = HEADER_FILL
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws_imp.row_dimensions[1].height = 30

    imp_headers = [
        "Clave trabajo",
        "Archivo vista",
        "Punto Venta",
        "Nro Comprobante",
        "Fecha",
        "Nombre Impuesto",
        "Porcentaje %",
        "Importe",
        "Estado Pago",
    ]
    imp_widths = [28, 36, 12, 14, 14, 35, 14, 16, 12]

    for col_idx, (header, width) in enumerate(zip(imp_headers, imp_widths), 1):
        cell = ws_imp.cell(row=2, column=col_idx)
        style_header_cell(cell, header, size=10)
        ws_imp.column_dimensions[get_column_letter(col_idx)].width = width
    ws_imp.row_dimensions[2].height = 35

    imp_row = 3
    for data in all_data:
        if data.get("_error"):
            continue
        cab = data.get("cabecera", {})
        impuestos = data.get("impuestos", []) or []
        estado = data.get("_estado_pago", "")
        job_key = data.get("_job_key", "")
        vista = data.get("_archivo", "")
        for i, imp in enumerate(impuestos):
            alt = (imp_row + i) % 2 == 0
            row_data = [
                job_key,
                vista,
                cab.get("punto_venta", ""),
                cab.get("numero_comprobante", ""),
                cab.get("fecha_emision", ""),
                imp.get("nombre", ""),
                imp.get("porcentaje", ""),
                imp.get("importe", ""),
                estado,
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_imp.cell(row=imp_row + i, column=col_idx)
                cell.value = value
                cell.font = Font(name="Arial", size=10)
                cell.fill = ALT_FILL if alt else WHITE_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = THIN_BORDER
            ws_imp.row_dimensions[imp_row + i].height = 18
        imp_row += max(len(impuestos), 1)

    ws_log = wb.create_sheet("Log de Errores")
    ws_log.sheet_view.showGridLines = False
    ws_log.merge_cells("A1:E1")
    t4 = ws_log["A1"]
    t4.value = "DATAFLOW — ERRORES DE PROCESAMIENTO"
    t4.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t4.fill = PatternFill("solid", start_color="C00000")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws_log.row_dimensions[1].height = 28

    for col_idx, header in enumerate(["Clave trabajo", "Archivo vista", "Error", "Timestamp", "Acción sugerida"], 1):
        cell = ws_log.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="C00000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws_log.column_dimensions["A"].width = 28
    ws_log.column_dimensions["B"].width = 36
    ws_log.column_dimensions["C"].width = 50
    ws_log.column_dimensions["D"].width = 20
    ws_log.column_dimensions["E"].width = 40

    log_row = 3
    errors_found = 0
    for data in all_data:
        if data.get("_error"):
            errors_found += 1
            for col_idx, value in enumerate(
                [
                    data.get("_job_key", ""),
                    data.get("_archivo", ""),
                    data.get("_error", ""),
                    data.get("_procesado", ""),
                    "Revisar PDF/página o reintentar",
                ],
                1,
            ):
                cell = ws_log.cell(row=log_row, column=col_idx)
                cell.value = value
                cell.font = Font(name="Arial", size=10)
                cell.fill = ERROR_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            ws_log.row_dimensions[log_row].height = 20
            log_row += 1

    if errors_found == 0:
        cell = ws_log.cell(row=3, column=1)
        cell.value = "✓ Sin errores — todos los trabajos procesados correctamente"
        cell.font = Font(name="Arial", size=11, bold=True, color="1F7A1F")
        ws_log.merge_cells("A3:E3")

    wb.save(output_path)
    return errors_found


def _try_load_repo_dotenv() -> None:
    """Carga .env / .env.local de la raíz del proyecto Dataflow (tres niveles arriba de este script)."""
    try:
        from dotenv import load_dotenv  # type: ignore
    except ImportError:
        return
    repo_root = Path(__file__).resolve().parent.parent.parent
    load_dotenv(repo_root / ".env")
    load_dotenv(repo_root / ".env.local")


def main():
    _try_load_repo_dotenv()

    parser = argparse.ArgumentParser(description="Parser facturas → Excel (DataFlow)")
    parser.add_argument("--input", required=True, help="Carpeta, archivo o glob")
    parser.add_argument("--output", default="facturas_dataflow.xlsx", help="Excel de salida")
    parser.add_argument("--delay", type=float, default=1.0, help="Segundos entre llamadas API")
    parser.add_argument(
        "--max-jobs",
        type=int,
        default=0,
        dest="max_jobs",
        help="Solo procesar los primeros N trabajos (páginas/imágenes); 0 = todos",
    )
    parser.add_argument(
        "--resume-from",
        type=int,
        default=0,
        dest="resume_from",
        help="Saltear los primeros N trabajos (página/imagen), no archivos sueltos",
    )

    args = parser.parse_args()

    files = get_input_files(args.input)
    if not files:
        print(f"ERROR: No hay archivos en: {args.input}")
        sys.exit(1)

    pdf_count = sum(1 for f in files if Path(f).suffix.lower() == ".pdf")
    if pdf_count and not HAS_PYMUPDF:
        print("ERROR: Hay PDF pero falta PyMuPDF.\n  pip install pymupdf")
        sys.exit(1)

    api_key = __import__("os").environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        print(
            "\nERROR: Falta ANTHROPIC_API_KEY (API de Anthropic / Claude).\n"
            "  Agregala en el archivo .env de la raíz del proyecto Dataflow, por ejemplo:\n"
            "    ANTHROPIC_API_KEY=sk-ant-api03-...\n"
            "  Reiniciá la terminal y volvé a ejecutar este script.\n"
            "  Documentación: INSTRUCCIONES_PARA_EL_DEV.md\n",
            file=sys.stderr,
        )
        sys.exit(2)

    jobs = flatten_jobs(files)
    if args.resume_from > 0:
        jobs = jobs[args.resume_from :]
        print(f"Retomando desde trabajo #{args.resume_from + 1} (total omitidos: {args.resume_from})")

    if args.max_jobs > 0:
        jobs = jobs[: args.max_jobs]
        print(f"Límite --max-jobs: procesando solo {len(jobs)} trabajo(s)")

    sep = "-" * 60
    print(f"\n{sep}")
    print("  DATAFLOW - PARSER DE FACTURAS")
    print(sep)
    print(f"  Archivos fuente: {len(files)}")
    print(f"  Trabajos (paginas/imagenes): {len(jobs)}")
    print(f"  Output: {args.output}")
    print(f"  Delay: {args.delay}s  |  modelo: {CLAUDE_MODEL}")
    print(f"{sep}\n")

    client = anthropic.Anthropic()
    all_data: list[dict] = []
    errors = 0

    for idx, job in enumerate(jobs, 1):
        label = job["vista_archivo"]
        print(f"[{idx:4d}/{len(jobs)}] {label} ...", end=" ", flush=True)

        try:
            if job["kind"] == "pdf":
                image_b64, media_type = encode_pdf_page_as_png(job["path"], job["page_index"])
            else:
                image_b64, media_type = encode_image_file(job["path"])

            data = extract_invoice_data(client, image_b64=image_b64, media_type=media_type, meta=job)
            all_data.append(data)

            cab = data.get("cabecera", {})
            items_count = len(data.get("items", []) or [])
            tipo = cab.get("tipo_comprobante", "?")
            proveedor = cab.get("razon_social", cab.get("proveedor_nombre_comercial", "?"))[:28]
            total = data.get("totales", {}).get("total_factura", "?")
            print(f"[OK] {tipo} | {proveedor} | ${total} | {items_count} items | {job['estado_pago']}")

        except json.JSONDecodeError as e:
            error_msg = f"JSON parse error: {str(e)}"
            print(f"[ERR] {error_msg[:50]}")
            all_data.append(
                {
                    "_job_key": job["job_key"],
                    "_archivo_pdf": job["pdf_name"],
                    "_pagina_pdf": job["page_display"],
                    "_estado_pago": job["estado_pago"],
                    "_archivo": job["vista_archivo"],
                    "_procesado": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "_error": error_msg,
                    "cabecera": {},
                    "items": [],
                    "impuestos": [],
                    "totales": {},
                    "notas": "",
                }
            )
            errors += 1

        except Exception as e:
            error_msg = str(e)
            print(f"[ERR] {error_msg[:70]}")
            all_data.append(
                {
                    "_job_key": job["job_key"],
                    "_archivo_pdf": job["pdf_name"],
                    "_pagina_pdf": job["page_display"],
                    "_estado_pago": job["estado_pago"],
                    "_archivo": job["vista_archivo"],
                    "_procesado": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "_error": error_msg,
                    "cabecera": {},
                    "items": [],
                    "impuestos": [],
                    "totales": {},
                    "notas": "",
                }
            )
            errors += 1

        if idx % 10 == 0:
            partial_path = args.output.replace(".xlsx", f"_parcial_{idx}.xlsx")
            create_excel(all_data, partial_path)
            print(f"  >> Guardado parcial: {partial_path}")

        if idx < len(jobs):
            time.sleep(args.delay)

    sep2 = "-" * 60
    print(f"\n{sep2}")
    print("  Generando Excel final...")
    create_excel(all_data, args.output)
    print(f"  OK archivo: {args.output}")
    print(f"  OK: {len(all_data) - errors}/{len(all_data)}  |  errores: {errors}")
    print(f"{sep2}\n")


if __name__ == "__main__":
    main()
